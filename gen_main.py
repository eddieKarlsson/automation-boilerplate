from re import I
import sys
import os
from os import listdir
from os.path import isfile, join
import json
import openpyxl as xl
from settings import Settings
from obj_lib.valve import Valve
from obj_lib.motor import Motor
from obj_lib.di import DI
from obj_lib.do import DO
from obj_lib.ai import AI
from obj_lib.ao import AO
from obj_lib.pid import PID
from obj_lib.sum import SUM
from obj_lib.alarm import Alarm
from obj_lib.asi import ASi
from unit_lib.unit_types import UnitTypes
from unit_lib.units_phases import UnitsPhases

class GenMain:
    """Main class called from UI,
    read data from excel and then execute sub-class functions
    """
    def __init__(self, excel_path, output_path, config_path):
        self.excel_path = excel_path
        self.output_path = output_path
        self.cm_output_path = os.path.join(self.output_path, 'CMs')
        self.alarm_output_path = os.path.join(self.output_path, 'Alarm')
        self.unit_output_path = os.path.join(self.output_path, 'Units_Phases')
        self.config_path = config_path        
        self.plcinexcel = set()
        self.s = Settings()
        self.dict_list = []
        self.tag_instance_counter = 0
        self.tia_tag_offset = 1000
        self.tia_tag_offset_add = 1000
        self.generate()

    def _open_gen_excel(self):
        try:
            wb = xl.load_workbook(self.excel_path, data_only=True)
        except FileNotFoundError as e:
            print(e)
            print('ERROR! Excel file not found, program will exit')
            sys.exit()
        else:
            self.wb = wb

    def copy_excel_data_to_dictionaries(self):
        """Open excel and read all relevant object-data to dict"""
        self._open_gen_excel()

        # Create all dictionaries, if enabled in settings
        if not self.s.DI_DISABLE:
            self.di_dict = self._obj_data_to_dict(
                        self.s.DI_SHEETNAME, self.s.DI_START_INDEX, 'di', tag=True)
            self.dict_list.append(self.di_dict)

        if not self.s.DO_DISABLE:
            self.do_dict = self._obj_data_to_dict(
                        self.s.DO_SHEETNAME, self.s.DO_START_INDEX, 'do', tag=True)
            self.dict_list.append(self.do_dict)

        if not self.s.VALVE_DISABLE:
            self.valve_dict = self._obj_data_to_dict(
                self.s.VALVE_SHEETNAME, self.s.VALVE_START_INDEX, 'valve', config=True)
            self.dict_list.append(self.valve_dict)

        if not self.s.MOTOR_DISABLE:
            self.motor_dict = self._obj_data_to_dict(
                self.s.MOTOR_SHEETNAME, self.s.MOTOR_START_INDEX, 'motor', config=True)
            self.dict_list.append(self.motor_dict)

        if not self.s.AI_DISABLE:
            self.ai_dict = self._obj_data_to_dict(
                self.s.AI_SHEETNAME, self.s.AI_START_INDEX, 'ai', eng_var=True, tag=True)
            self.dict_list.append(self.ai_dict)

        if not self.s.AO_DISABLE:
            self.ao_dict = self._obj_data_to_dict(
                    self.s.AO_SHEETNAME, self.s.AO_START_INDEX, 'ao', eng_var=True, tag=True)
            self.dict_list.append(self.ao_dict)

        if not self.s.PID_DISABLE:
            self.pid_dict = self._obj_data_to_dict(
                    self.s.PID_SHEETNAME, self.s.PID_START_INDEX, 'pid', eng_var=True)
            self.dict_list.append(self.pid_dict)

        if not self.s.SUM_DISABLE:
            self.sum_dict = self._obj_data_to_dict(
                    self.s.SUM_SHEETNAME, self.s.SUM_START_INDEX, 'sum', eng_var=True, volumeperpulse=True)
            self.dict_list.append(self.sum_dict)

        if not self.s.ALARM_DISABLE:
            self.alarm_dict = self._obj_data_to_dict(
                    self.s.ALARM_SHEETNAME, self.s.ALARM_START_INDEX, 'alarm', generic_alarm=True)
            self.dict_list.append(self.alarm_dict)

        if not self.s.ASI_DISABLE:
            self.asi_dict = self._obj_data_to_dict(
                    self.s.ASI_SHEETNAME, self.s.ASI_START_INDEX, 'asi', asi=True)
            self.dict_list.append(self.asi_dict)
        
        if not self.s.UNIT_DISABLE:
            self.unit_phase_list = self._unit_data_to_list(self.s.UNIT_SHEETNAME)

    def _obj_data_to_dict(self, sheet, start_index, type, config=False, eng_var=False, volumeperpulse=False,
                          generic_alarm=False, asi=False, tag=False):
        """Read all object data to dict"""

        # Open excel sheet
        try:
            ws = self.wb[sheet]
        except KeyError:
            msg = f'ERROR! {sheet} sheet does not exist, program will exit'
            print(msg)
            sys.exit()

            # Loop header and set the corresponding variables to
            # the integer number
        for i in range(1, 20):
            cell = ws.cell(row=self.s.HEADER_ROW, column=i)
            cellval = str(cell.value)

            # If cell is empty (NoneType) - skip it
            if cellval is None:
                continue

            if self.s.COL_ID_NAME == cellval:
                column_id = i
            if self.s.COL_COMMENT_NAME == cellval:
                column_comment = i
            if self.s.COL_ALARM_GROUP_NAME == cellval:
                column_alarmgroup = i
            if self.s.COL_PLC_NAME == cellval:
                column_plc = i

            if config:
                if self.s.COL_CONFIG_NAME == cellval:
                    column_config = i

            if volumeperpulse:
                if self.s.COL_VolumePerPulse_Name == cellval:
                    column_volumeperpulse = i

            if eng_var:
                if self.s.COL_ENG_UNIT_NAME == cellval:
                    column_eng_unit = i
                if self.s.COL_ENG_MIN_NAME == cellval:
                    column_eng_min = i
                if self.s.COL_ENG_MAX_NAME == cellval:
                    column_eng_max = i

            if generic_alarm:
                if self.s.COL_ALARM_PRIO_NAME == cellval:
                    column_alarm_prio = i
                if self.s.COL_ALARM_TEXT_NAME == cellval:
                    column_alarm_text = i

            if asi:
                if self.s.COL_ASI_ADDR_NAME == cellval:
                    column_asi_addr = i
                if self.s.COL_ASI_MASTER_NAME == cellval:
                    column_asi_master = i

        if self.s.debug_level > 0:
            print('SHEET:', sheet)
            print('\t', 'column_id:', column_id)
            print('\t', 'column_comment:', column_comment)
            print('\t', 'column_alarmgroup:', column_alarmgroup)
            print('\t', 'column_plc:', column_plc)
            if config:
                print('\t', 'column_config:', column_config)
            if volumeperpulse:
                print('\t', 'column_volumeperpulse:', column_volumeperpulse)
            if eng_var:
                print('\t', 'column_eng_unit:', column_eng_unit)
                print('\t', 'column_eng_min:', column_eng_min)
                print('\t', 'column_eng_max:', column_eng_max)
            if generic_alarm:
                print('\t', 'column_alarm_prio:', column_alarm_prio)
                print('\t', 'column_alarm_text:', column_alarm_text)
            if asi:
                print('\t', 'column_asi_addr:', column_asi_addr)
                print('\t', 'column_asi_master:', column_asi_master)

        # Handle tag offsets        
        if tag:
            self.tag_instance_counter += 1
            if self.tag_instance_counter > 1:
                self.tia_tag_offset += self.tia_tag_offset_add
            self.create_tia_memory(start_address=self.tia_tag_offset, initialize=True)
            

        # Loop through object list and add key-value pairs to object dict
        # then append each object-dict to list
        obj_list = []
        index = start_index
        for i in range(self.s.ROW, ws.max_row + 1):
            # Break if we get a blank ID cell
            cell_id = ws.cell(row=i, column=column_id)
            cell_comment = ws.cell(row=i, column=column_comment)
            cell_alarmgroup = ws.cell(row=i, column=column_alarmgroup)
            cell_plc = ws.cell(row=i, column=column_plc)

            if cell_id.value is None:
                break

            # Always insert these key-value pairs
            obj = {
                'type': type,
                'id': cell_id.value,
                'comment': cell_comment.value,
                'index': index,
                'alarmgroup': cell_alarmgroup.value,
                'plc': cell_plc.value,
            }

            # Add conditional key-value pairs
            if config:
                cell_config = ws.cell(row=i, column=column_config)
                obj['config'] = cell_config.value

            if volumeperpulse:
                cell_volumeperpulse = ws.cell(row=i, column=column_volumeperpulse)
                obj['volumeperpulse'] = cell_volumeperpulse.value

            if eng_var:
                cell_eng_unit = ws.cell(row=i, column=column_eng_unit)
                obj['eng_unit'] = cell_eng_unit.value
                cell_eng_min = ws.cell(row=i, column=column_eng_min)
                obj['eng_min'] = cell_eng_min.value
                cell_eng_max = ws.cell(row=i, column=column_eng_max)
                obj['eng_max'] = cell_eng_max.value

            if generic_alarm:
                cell_alarm_text = ws.cell(row=i, column=column_alarm_text)
                obj['alarm_text'] = cell_alarm_text.value
                cell_alarm_prio = ws.cell(row=i, column=column_alarm_prio)
                obj['alarm_prio'] = cell_alarm_prio.value

            if asi:
                cell_asi_addr = ws.cell(row=i, column=column_asi_addr)
                obj['asi_addr'] = cell_asi_addr.value
                cell_asi_master = ws.cell(row=i, column=column_asi_master)
                obj['asi_master'] = cell_asi_master.value

            if obj['type'] == 'ai' or obj['type'] == 'ao' or obj['type'] == 'sum':
                tmp_mem_size = 2
            else:
                tmp_mem_size = 0
            
            if tag:
                obj['tag'] = self.create_tia_memory(memory_size_byte=tmp_mem_size)
            
            obj_list.append(obj)
            index += 1

        for obj in obj_list:
            self.plcinexcel.add(obj['plc'])

        return obj_list

    def _unit_data_to_list(self, sheet):
        """Read all unit object data to dict"""

        # Open excel sheet
        try:
            ws = self.wb[sheet]
        except KeyError:
            msg = f'ERROR! {sheet} sheet does not exist, program will exit'
            print(msg)
            sys.exit()
        
            column_db_start_addr = None
            # Loop header and set the corresponding variables to
            # the integer number
        for i in range(1, 10):
            cell = ws.cell(row=self.s.UNIT_HEADER_ROW, column=i)
            cellval = str(cell.value)

            # If cell is empty (NoneType) - skip it
            if cellval is None:
                continue

            if self.s.COL_ID_NAME == cellval:
                column_id = i
            elif self.s.COL_TYPE_NAME == cellval:
                column_type = i            
            elif self.s.COL_PLC_NAME == cellval:
                column_plc = i
            elif self.s.COL_ALARM_GROUP_NAME == cellval:
                column_hmi_group = i
            elif self.s.COL_DB_START_ADDR_NAME == cellval:
                column_db_start_addr = i

        if self.s.debug_level > 0:
            print('UNIT UNITSHEET:', sheet)
            print('\t', 'UNIT column_id:', column_id)
            print('\t', 'UNIT column_type:', column_type)
            print('\t', 'UNIT column_plc:', column_plc)
            print('\t', 'UNIT column_hmi_group:', column_hmi_group)
            print('\t', 'UNIT column_db_start_addr:', column_db_start_addr)


        unit_phase_list = []
        #  loop over the objects in sheet
        mem_unit = None
        mem_plc = None

        if column_db_start_addr is not None:
            db_addr_exists = True
        else:
            db_addr_exists = False

        for i in range(self.s.UNIT_ROW, ws.max_row + 1):
            #  Create cell references
            cell_id = ws.cell(row=i, column=column_id)
            cell_type = ws.cell(row=i, column=column_type)
            cell_plc = ws.cell(row=i, column=column_plc)
            cell_hmi_group = ws.cell(row=i, column=column_hmi_group)
            if db_addr_exists:
                cell_db_start_addr = ws.cell(row=i, column=column_db_start_addr)


            # Break if we get a blank ID cell
            if cell_id.value is None:
                break
            
            def _is_valid_unit_type(in_type):
                for type in UnitTypes:
                    if in_type == type.value:
                        return True
                return False

            def _is_unit(in_type):
                if "Unit" in in_type:
                    return True
                else:
                    return False

            is_valid_unit_type = _is_valid_unit_type(cell_type.value)
            is_unit = _is_unit(cell_type.value)
            is_phase = not is_unit

            if is_unit:
                #  Remembers var from unit, in that way less
                #  duplicate data in excel
                mem_unit = cell_id.value
                mem_plc = cell_plc.value
                mem_hmi_group = cell_hmi_group.value
                parent = None
            else:
                parent = mem_unit

            # Create object dict, always with these key-value pairs
            obj = {
                'is_valid_unit_type': is_valid_unit_type,
                'is_unit': is_unit,
                'is_phase': is_phase,
                'parent': parent,
                'type': cell_type.value,
                'id': cell_id.value,
                'plc': mem_plc,
                'hmi_group': mem_hmi_group,
            }

            # Insert DB start addr if property exists
            if db_addr_exists:
                db, db_offset = self._parse_s7_db_addr(cell_db_start_addr.value)
                obj['db_nr_str'] = db
                obj['db_offset'] = db_offset

            unit_phase_list.append(obj)
        
        return unit_phase_list

    def create_subdirs(self):
        """Create all subdirectiories beyond output path"""
        dirs = [self.s.TIA_DIR, self.s.INTOUCH_DIR, self.s.SQL_DIR]
        for dir in dirs:
            newdir = os.path.join(self.output_path, dir)
            if not os.path.exists(newdir):
                os.makedirs(newdir)

    def create_subdirsplc(self):
        """Create all subdirectiories beyond output path"""
        for plc in self.plcinexcel:
            newdir = os.path.join(self.output_path, self.s.TIA_DIR)
            newdir = os.path.join(newdir, plc)
            if not os.path.exists(newdir):
                os.makedirs(newdir)

    def get_config_from_config_path(self):
        """Load config from .json file"""
        json_file = os.path.join(self.config_path, 'config_type.json')
        with open(json_file, 'r') as (f):
            json_var = json.load(f)
            self.config_type = json_var['type']
            print(f'Config Type={self.config_type}')

    @staticmethod
    def _print_disabled_in_settings(prefix):
        print(f"{prefix} not generated, disabled in settings file")

    def _combine_it_files(self):
        """
        combines all it files in folder and chops off first lines
        in files other than the first
        """

        for folder in listdir(self.output_path):
            filename = "ALL_" + folder + "_IT.csv"
            path = os.path.join(self.output_path, folder, self.s.INTOUCH_DIR)
            outfile = os.path.join(path, filename)

            if os.path.exists(outfile):
                os.remove(outfile)

            if os.path.exists(path):
                file_list = [f for f in listdir(path)
                    if isfile(join(path, f))]            

                with open(outfile, 'w', encoding='cp1252') as wf:
                        for file_index, file in enumerate(file_list):
                            with open(os.path.join(path, file), 'r', encoding='cp1252') as rf:
                                for line_index, line in enumerate(rf):
                                    # Skip first line header if it's not the first file
                                    if file_index > 0 and line_index <= 0:
                                            continue
                                    wf.write(line)                    

    def _combine_sql_files(self):
        """
        combines all it files in folder and chops off first lines
        in files other than the first
        """

        for folder in listdir(self.output_path):
            filename = "ALL_" + folder + "_SQL.csv"
            path = os.path.join(self.output_path, folder, self.s.SQL_DIR)
            outfile = os.path.join(path, filename)

            if os.path.exists(outfile):
                os.remove(outfile)

            if os.path.exists(path):
                file_list = [f for f in listdir(path)
                    if isfile(join(path, f))]            

                with open(outfile, 'w', encoding='cp1252') as wf:
                        for file_index, file in enumerate(file_list):
                            with open(os.path.join(path, file), 'r', encoding='cp1252') as rf:
                                for line_index, line in enumerate(rf):
                                    # Skip first line header if it's not the first file
                                    if file_index > 0 and line_index <= 0:
                                            continue
                                    wf.write(line)   

    @staticmethod
    def _parse_s7_db_addr(in_db_addr):
        #  Format expected e.g. DB9001.DBX12.0, DB9001.DBB12, DB9001.DBW12, DB9001.DBD12
        splitted = in_db_addr.split('.')

        if len(splitted) == 3: 
            # three elements == two dots == DB9001.DBX0.0 format
            db, dbx, bit = splitted
        elif len(splitted) == 2:
            #  two elements == one dot == any other format specified
            db, dbx = splitted

        remove_chars = ['D', 'B', 'X', 'W', 'D']

        for char in remove_chars:
            dbx = dbx.replace(char, '')

        db_offset = int(dbx)

        return db, db_offset

    def generate(self):
        print('Version', self.s.version)

        self.copy_excel_data_to_dictionaries()

        if self.s.debug_level > 0:
            for dict in self.dict_list:
                for obj in dict:
                    print(obj)

        if not self.s.UNIT_DISABLE and self.s.debug_level > 0:
            for _ in self.unit_phase_list:
                print(_)

        self.get_config_from_config_path()

        if self.s.VALVE_DISABLE:
            self._print_disabled_in_settings('Valve')
        else:
            Valve(self, self.output_path, self.valve_dict, self.config_path,
                  config_type=self.config_type)

        if self.s.MOTOR_DISABLE:
            self._print_disabled_in_settings('Motor')
        else:
            Motor(self, self.output_path, self.motor_dict, self.config_path,
                  config_type=self.config_type)

        if self.s.DI_DISABLE:
            self._print_disabled_in_settings('DI')
        else:
            DI(self, self.output_path, self.di_dict, self.config_path,
                config_type=self.config_type)

        if self.s.DO_DISABLE:
            self._print_disabled_in_settings('DO')
        else:
            DO(self, self.output_path, self.do_dict, self.config_path,
                config_type=self.config_type)

        if self.s.AI_DISABLE:
            self._print_disabled_in_settings('AI')
        else:
            AI(self, self.output_path, self.ai_dict, self.config_path,
                config_type=self.config_type)

        if self.s.AO_DISABLE:
            self._print_disabled_in_settings('AO')
        else:
            AO(self, self.output_path, self.ao_dict, self.config_path,
                config_type=self.config_type)

        if self.s.PID_DISABLE:
            self._print_disabled_in_settings('PID')
        else:
            PID(self, self.output_path, self.pid_dict, self.config_path,
                config_type=self.config_type)

        if self.s.SUM_DISABLE:
            self._print_disabled_in_settings('SUM')
        else:
            SUM(self, self.output_path, self.sum_dict, self.config_path,
                config_type=self.config_type)

        if self.s.ALARM_DISABLE:
            self._print_disabled_in_settings('Alarm')
        else:
            Alarm(self, self.output_path, self.alarm_dict, self.config_path,
                config_type=self.config_type)

        if self.s.ASI_DISABLE:
            self._print_disabled_in_settings('ASi')
        else:
            ASi(self, self.output_path, self.asi_dict, self.config_path, config_type=self.config_type)


        if self.s.UNIT_DISABLE:
            self._print_disabled_in_settings('Units_Phases')
        else: 
            UnitsPhases(self, self.output_path, self.unit_phase_list, 
                          self.config_path, config_type=self.config_type)

        self._combine_it_files()
        self._combine_sql_files()

    def create_tia_memory(self, start_address=0, initialize=False, memory_size_byte=2):
        """Returns a memory unique memory address by counting up"""
        if initialize:
            self.tia_bit = -1
            self.tia_byte = start_address            
            return
        
        if memory_size_byte == 0:  # bit
            self.tia_bit += 1
            if self.tia_bit > 7:
                self.tia_bit = 0
                self.tia_byte += 1
            return f"M{self.tia_byte}.{self.tia_bit}"
        elif memory_size_byte == 4:  # Dword
            self.tia_byte += memory_size_byte
            return f"MD{self.tia_byte}"
        else:  # If not bit or Dword always presume its a 2 byte Int
            self.tia_byte += 2  # Word
            return f"MW{self.tia_byte}"
